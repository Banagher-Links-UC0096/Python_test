import serial
import serial.tools.list_ports
import time
from datetime import datetime
import struct
import tkinter as tk
from tkinter import ttk
import openpyxl
from openpyxl import Workbook
import os
import threading

modbus_data = {
    0x000b:[1, 1,'Product type',0,["Controller","Controller","Inverter","Integrated inverter controller","Main friequency off-grid"]],
    
    0x0100:[1, 5,'Battery level SOC(%)',1,"%"],
    0x0101:[1, 3,'Battery voltage(V)',10,"V"],
    0x0102:[1, 4,'Battery current(A)',10,"A"],

    0x0107:[1,12,'PV voltage(V)',10,"V"],
    0x0108:[1,13,'PV charging current(A)',10,"A"],
    0x0109:[1,14,'PV charging power(W)',1,"W"],0x010a:[1,0],
    0x010b:[1, 0,'Charge state',0,["Not start","Const Current","Const Voltage","-","Float","-","Active","Active"]],0x010c:[1,0],0x010d:[1,0],
    0x010e:[1,16,'Charging power(W)',1,"W"],0x010f:[1,0],

    0x0204:[1,0],0x0205:[1,0],0x0206:[1,0],0x0207:[1,0],0x0208:[1,0],0x0209:[1,0],0x020a:[1,0],0x020b:[1,0],0x020c:[1,0],0x020d:[1,0],0x020e:[1,0],0x020f:[1,0],
    0x0210:[1, 2,'Current state of machine',0,
            ["Power-on","Stand by","Intialization","Soft start","Running in line","Running in invert","Invert to line","Line to invert","remain","remain","Shutdown","Fault"]],0x0211:[0],

    0x0212:[1, 7,'Bus voltage(V)',10,"V"],
    0x0213:[1, 8,'Mains voltage(V)',10,"V"],
    0x0214:[1, 9,'Grid current(A)',10,"A"],
    0x0215:[1,10,'Mains frequency(Hz)',100,"Hz"],
    0x0216:[1,17,'Output voltage(V)',10,"V"],
    0x0217:[1,18,'InvCurr(A)',10],
    0x0218:[1,19,'Output frequency(Hz)',100,"Hz"],
    0x0219:[1,20,'Load current(A)',10,"A"],0x021a:[1,0],
    0x021b:[1,21,'Load active power(W)',1,"W"],
    0x021c:[1,22,'Apparent power of load(VA)',1,"VA"],0x021d:[1,0],
    0x021e:[1,11,'Mains charging current(A)',10,"A"],
    0x021f:[1,23,'Load rate(%)',1,"%"],
    0x0220:[1,24,'PV radiator temperature(°C)',10,"°C"],
    0x0221:[1,25,'Temperature of inverter heat sink(°C)',10,"°C"],
    0x0222:[1,26,'Inverter radiator temperature(°C)',10,"°C"],
    0x0223:[1, 0,'Ambient temperature(°C)',10,"°C"],
    0x0224:[1,15,'PV charging current(A)',10,"A"],
    0x0225:[1, 0,'Back current(A)',10,"A"],
    
    0xe004:[1, 6,'BatTypeSet',0,["User-defined","SLD","FLD","GEL","LFPx14","LFPx15","LFPx16","LFPx7","LFPx8","LFPx9","NCAx7","NCAx8","NCAx13","NCAx14"]],

    0xe116:[1,38,'Equipment type',0,["","","","","","","","","","","","","","HF2430U60-100","","","","","","","","HF2430S80-H","","","","","","","","","","","","","HYP4850U100-H"]],

    0xf000:[1,0],0xf001:[1,0],0xf002:[1,0],0xf003:[1,0],0xf004:[1,0],0xf005:[1,0],0xf006:[1,0],
    0xf007:[1,0],0xf008:[1,0],0xf009:[1,0],0xf00a:[1,0],0xf00b:[1,0],0xf00c:[1,0],0xf00d:[1,0],
    0xf00e:[1,0],0xf00f:[1,0],0xf010:[1,0],0xf011:[1,0],0xf012:[1,0],0xf013:[1,0],0xf014:[1,0],
    0xf015:[1,0],0xf016:[1,0],0xf017:[1,0],0xf018:[1,0],0xf019:[1,0],0xf01a:[1,0],0xf01b:[1,0],
    0xf01c:[1,0],0xf01d:[1,0],0xf01e:[1,0],0xf01f:[1,0],0xf020:[1,0],0xf021:[1,0],0xf022:[1,0],
    0xf023:[1,0],0xf024:[1,0],0xf025:[1,0],0xf026:[1,0],0xf027:[1,0],0xf028:[1,0],0xf029:[1,0],

    0xf02d:[1,27,'Ampere-hours of battery charging on the same day(AH)',1,"AH"],
    0xf02e:[1,28,'Ampere-hours of battery discharge on the same day(AH)',1,"AH"],
    0xf02f:[1,29,'PV daily power generation(kWh)',10,"kWh"],
    0xf030:[1,30,'Electricity consumption on the day of load(kWh)',10,"kWh"],0xf031:[3,0],

    0xf034:[2,32,'Accumulated battery charging ampere hours(AH)',1,"AH"],
    0xf036:[2,33,'Accumulated discharge ampere hours of battery(AH)',1,"AH"],
    0xf038:[2,35,'PV cumulative power generation(kWh)',10,"kWh"],
    0xf03a:[2,36,'Accumulated power consumption of load(kWh)',10,"kWh"],
    0xf03c:[1, 0,'Mains charge level of the day(AH)',1,"AH"],
    0xf03d:[2,31,'Consumption of municipal electricity on the day of load(kWh)',10,"kWh"],

    0xf040:[3,0],0xf043:[3,0],
    0xf046:[1,34,'Accumulated charging capacity of municipal electricity(AH)',10,"AH"],
    0xf048:[1,37,'Accumulated load from mains consumption(kWh)',10,"kWh"],
    0xf04a:[1, 0,'Accumulated working hoursof inverter(H)',1,"H"],
    0xf04b:[1, 0,'Aooumulated working hoursof bypass(H)',1,"H"],
    }

# --- ここでグローバルに定義 ---
modbus_items = [v for v in modbus_data.values() if len(v) > 2 and isinstance(v[1], int) and v[1] > 0]
modbus_items_sorted = sorted(modbus_items, key=lambda x: x[1])
modbus_addr_sorted = [
    addr for addr, v in sorted(
        ((addr, v) for addr, v in modbus_data.items() if len(v) > 2 and isinstance(v[1], int) and v[1] > 0),
        key=lambda x: x[1][1]
    )
]
# --- ポートの選択と初期化 ---
def select_com_ports():
    """
    使用可能なCOMポートを選択するウィンドウを画面中央に表示し、Logger1のデフォルトをCOM10、Inverter1のデフォルトをCOM13にする。
    """
    ports = [port.device for port in serial.tools.list_ports.comports()]
    if not ports:
        print("No COM ports available.")
        return None, None, None, None  # 全てがNoneの場合

    selected_ports = {"Logger_port1": None, "Logger_port2": None, "Inverter_port1": None, "Inverter_port2": None}

    def update_com_options(*args):
        selected_logger1 = logger1_var.get()
        selected_logger2 = logger2_var.get()
        selected_inverter1 = inverter1_var.get()
        selected_inverter2 = inverter2_var.get()
        logger1_dropdown['values'] = [port for port in ports if port not in {selected_logger2, selected_inverter1, selected_inverter2}]
        logger2_dropdown['values'] = ["なし"] + [port for port in ports if port not in {selected_logger1, selected_inverter1, selected_inverter2}]
        inverter1_dropdown['values'] = ["なし"] + [port for port in ports if port not in {selected_logger1, selected_logger2, selected_inverter2}]
        inverter2_dropdown['values'] = ["なし"] + [port for port in ports if port not in {selected_logger1, selected_logger2, selected_inverter1}]

    def on_submit():
        selected_ports["Logger_port1"] = logger1_var.get()
        selected_ports["Logger_port2"] = logger2_var.get()
        selected_ports["Inverter_port1"] = inverter1_var.get()
        selected_ports["Inverter_port2"] = inverter2_var.get()
        root.destroy()

    root = tk.Tk()
    root.title("USBポートの選択")

    # 画面中央に配置
    root.update_idletasks()
    width = 350
    height = 200
    x = (root.winfo_screenwidth() // 2) - (width // 2)
    y = (root.winfo_screenheight() // 2) - (height // 2)
    root.geometry(f"{width}x{height}+{x}+{y}")

    # Logger_portの選択
    tk.Label(root, text="Logger1 Port:", anchor="w").grid(row=0, column=0, padx=10, pady=5)
    logger1_default = "COM10" if "COM10" in ports else (ports[0] if ports else "なし")
    logger1_var = tk.StringVar(value=logger1_default)
    logger1_var.trace("w", update_com_options)
    logger1_dropdown = ttk.Combobox(root, textvariable=logger1_var, values=ports, state="readonly")
    logger1_dropdown.grid(row=0, column=1, padx=10, pady=5)

    tk.Label(root, text="Logger2 Port:", anchor="w").grid(row=1, column=0, padx=10, pady=5)
    logger2_var = tk.StringVar(value="なし")
    logger2_var.trace("w", update_com_options)
    logger2_dropdown = ttk.Combobox(root, textvariable=logger2_var, values=["なし"] + ports, state="readonly")
    logger2_dropdown.grid(row=1, column=1, padx=10, pady=5)

    # Inverter_portの選択
    tk.Label(root, text="Inverter1 Port:", anchor="w").grid(row=2, column=0, padx=10, pady=5)
    inverter1_default = "COM13" if "COM13" in ports else "なし"
    inverter1_var = tk.StringVar(value=inverter1_default)
    inverter1_var.trace("w", update_com_options)
    inverter1_dropdown = ttk.Combobox(root, textvariable=inverter1_var, values=["なし"] + ports, state="readonly")
    inverter1_dropdown.grid(row=2, column=1, padx=10, pady=5)

    tk.Label(root, text="Inverter2 Port:", anchor="w").grid(row=3, column=0, padx=10, pady=5)
    inverter2_var = tk.StringVar(value="なし")
    inverter2_var.trace("w", update_com_options)
    inverter2_dropdown = ttk.Combobox(root, textvariable=inverter2_var, values=["なし"] + ports, state="readonly")
    inverter2_dropdown.grid(row=3, column=1, padx=10, pady=5)

    # 確定ボタン
    submit_button = tk.Button(root, text="設定", command=on_submit)
    submit_button.grid(row=4, column=0, columnspan=2, pady=10)

    root.mainloop()

    return (
        selected_ports["Logger_port1"] if selected_ports["Logger_port1"] != "なし" else None,
        selected_ports["Logger_port2"] if selected_ports["Logger_port2"] != "なし" else None,
        selected_ports["Inverter_port1"] if selected_ports["Inverter_port1"] != "なし" else None,
        selected_ports["Inverter_port2"] if selected_ports["Inverter_port2"] != "なし" else None,
    )

# COMポートの設定
LOGGER_PORT1, LOGGER_PORT2, INVERTER_PORT1, INVERTER_PORT2 = select_com_ports()
if LOGGER_PORT1 or LOGGER_PORT2 or INVERTER_PORT1 or INVERTER_PORT2:
    print(f"Logger Port 1: {LOGGER_PORT1}, Logger Port 2: {LOGGER_PORT2}, Inverter Port 1: {INVERTER_PORT1}, Inverter Port 2: {INVERTER_PORT2}")
else:
    print("No valid ports selected.")
    exit()

BAUD_RATE = 9600  # ボーレートは必要に応じて変更
BYTE_SIZE = 8  # データビット数
PARITY = "N"  # パリティビット
STOP_BITS = 1  # ストップビット数
TIMEOUT = 1  # タイムアウト時間（秒）

# シリアルポートの初期化（インスタンスを分離）
try:
    ser_logger1 = serial.Serial(LOGGER_PORT1, BAUD_RATE, BYTE_SIZE, PARITY, STOP_BITS, TIMEOUT) if LOGGER_PORT1 else None
    ser_inverter1 = serial.Serial(INVERTER_PORT1, BAUD_RATE, BYTE_SIZE, PARITY, STOP_BITS, TIMEOUT) if INVERTER_PORT1 else None
    inverter1_connected = ser_inverter1 is not None

    ser_logger2 = serial.Serial(LOGGER_PORT2, BAUD_RATE, timeout=1) if LOGGER_PORT2 else None
    ser_inverter2 = serial.Serial(INVERTER_PORT2, BAUD_RATE, timeout=1) if INVERTER_PORT2 else None
    inverter2_connected = ser_inverter2 is not None

    logger2_connected = ser_logger2 is not None

except serial.SerialException:
    print("ERROR")
    exit()

# CRC16 計算関数
def calc_crc16(data):
    """
    Modbus RTU用CRC16計算
    """
    crc = 0xFFFF
    for byte in data:
        crc ^= byte
        for _ in range(8):
            if crc & 0x0001:
                crc = (crc >> 1) ^ 0xA001
            else:
                crc >>= 1
    return crc & 0xFFFF

# Modbusフレームからアドレスデータ抽出
def extract_modbus_register_address(data):
    """ Modbusフレームからレジスターアドレスを抽出し、データ数分のアドレスリストを作成 """
    if len(data) >= 6:  # Modbusフレームの最低長
        base_address = int.from_bytes(data[2:4], byteorder='big')  # レジスターアドレス（3,4バイト目）
        register_count = int.from_bytes(data[4:6], byteorder='big')  # レジスターデータの数（5バイト目）
        
        address_list = [hex(base_address + i) for i in range(register_count)]  # アドレスリスト作成
        
        #print(f"抽出したレジスターアドレス: {hex(base_address)}")
                
        return address_list
    return []

def extract_modbus_register_value(data):
    """
    Modbusフレームからレジスターデータを抽出し、
    modbus_dataのリスト1番目が1の場合は2バイト、2の場合は4バイト、3の場合は6バイトで1つの値として処理しリスト化する。
    CRCエラーの場合はデータを無効とする。
    """
    if len(data) >= 7:  # レスポンスフレームの最低長
        data_bytes = data[3:-2]  # データ数バイト以降がデータ部
        base_addr = int.from_bytes(data[2:4], byteorder='big') if len(data) >= 4 else 0
        data_list = []
        idx = 0
        addr = base_addr
        while idx < len(data_bytes):
            word_count = 1
            if addr in modbus_data and len(modbus_data[addr]) > 0:
                word_count = modbus_data[addr][0]
            if word_count == 3 and idx + 6 <= len(data_bytes):
                value = (
                    (int.from_bytes(data_bytes[idx:idx+2], byteorder='big', signed=False) << 32) |
                    (int.from_bytes(data_bytes[idx+2:idx+4], byteorder='big', signed=False) << 16) |
                    int.from_bytes(data_bytes[idx+4:idx+6], byteorder='big', signed=False)
                )
                idx += 6
            elif word_count == 2 and idx + 4 <= len(data_bytes):
                value = (
                    (int.from_bytes(data_bytes[idx:idx+2], byteorder='big', signed=False) << 16) |
                    int.from_bytes(data_bytes[idx+2:idx+4], byteorder='big', signed=False)
                )
                idx += 4
            elif word_count == 1 and idx + 2 <= len(data_bytes):
                scale = 1
                signed_flag = False
                if addr in modbus_data and len(modbus_data[addr]) > 3:
                    scale = modbus_data[addr][3]
                    # 4番目が負ならsignedで変換し、絶対値で割る
                    #if isinstance(scale, (int, float)) and scale < 0:
                    #    signed_flag = True
                    #    scale = abs(scale)
                if signed_flag:
                    # 2バイト値をsignedで取得
                    value = int.from_bytes(data_bytes[idx:idx+2], byteorder='big', signed=True)
                    value = value / scale
                else:
                    value = int.from_bytes(data_bytes[idx:idx+2], byteorder='big', signed=False)
                    if isinstance(scale, (int, float)) and scale > 1:
                        value = value / scale
                idx += 2
            else:
                break
            # 文字列変換
            if addr in modbus_data:
                v = modbus_data[addr]
                if len(v) > 4 and v[3] == 0 and isinstance(v[4], list):
                    try:
                        if 0 <= value < len(v[4]):
                            value = v[4][int(value)]
                    except Exception:
                        pass
            data_list.append(value)
            addr += 1
        return data_list
    return []

# Excelファイルに書き込む関数
def write_to_excel(log_data_list, file_name):
    """
    ログデータをExcelファイルに追記する。
    2行目を改行し、2行目にデータを書き込む（既存データは3行目以降にシフト）。
    """
    try:
        if os.path.exists(file_name):
            wb = openpyxl.load_workbook(file_name)
            ws = wb.active
            # 2行目に空行を挿入してデータを2行目に書き込む
            ws.insert_rows(2)
            for i, row in enumerate(log_data_list):
                for j, value in enumerate(row, 1):
                    ws.cell(row=2, column=j, value=value)
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Log Data"
            global modbus_items_sorted
            headers = ['Timestamp'] + [item[2] for item in modbus_items_sorted]
            headers += [
                'battery energy today discharge(Wh)',
                'battery energy today charge(Wh)',
                'battery energy total charge(Wh)',
                'battery energy total discharge(Wh)',
            ]
            ws.append(headers)
            for row in log_data_list:
                ws.append(row)
        print(log_data_list)
        wb.save(file_name)
    except Exception as e:
        print(f"Excelファイル書き込み中にエラーが発生しました: {e}")

def process_and_save_data(timestamp, address_list, value_list, file_name):
    """
    address_list: 受信したアドレスリスト（例: [0x000b, 0x0101, ...]）
    value_list:   受信した値リスト（例: [12, 34, ...]）
    4番目のデータと29番目、28番目、33番目、34番目のデータの積算も追加して保存
    積算値は小数点第1位で四捨五入
    """
    # --- 0x0102のデータが""の場合は何もしない ---
    if 0x0102 in address_list:
        idx_0102 = address_list.index(0x0102)
        val_0102 = value_list[idx_0102]
        if val_0102 != "":
            v = modbus_data.get(0x0102)
            if v and len(v) > 3:
                scale = v[3]
                if isinstance(val_0102, (int, float)) and val_0102 >= 32768 and isinstance(scale, (int, float)) and scale != 0:
                    val_signed = int(val_0102) - 65536
                    value_list[idx_0102] = val_signed #/ abs(scale)
                elif isinstance(scale, (int, float)) and scale < 0:
                    # scaleが負の場合はsigned変換
                    val_int = int(val_0102)
                    if val_int > 32767:
                        val_signed = val_int - 65536
                    else:
                        val_signed = val_int
                    value_list[idx_0102] = val_signed / abs(scale)
                elif isinstance(scale, (int, float)) and scale > 1:
                    value_list[idx_0102] = float(val_0102) #/ scale

    addr_val_dict = {addr: val for addr, val in zip(address_list, value_list)}
    row = [timestamp]
    for addr in modbus_addr_sorted:
        val = addr_val_dict.get(addr, "")
        # --- 通常のデータタイプ判定 ---
        if val == "":
            row.append("")
            continue
        v = modbus_data.get(addr)
        if v and len(v) > 3:
            data_type = v[3]
            if data_type == 0 and len(v) > 4 and isinstance(v[4], list):
                try:
                    idx = int(val)
                    if 0 <= idx < len(v[4]):
                        row.append(v[4][idx])
                    else:
                        row.append(val)
                except Exception:
                    row.append(val)
            elif data_type in (10, 100, 1000):
                try:
                    row.append(float(val) / data_type)
                except Exception:
                    row.append(val)
            else:
                row.append(val)
        else:
            row.append(val)

    # 積算値の追加（小数点第1位で四捨五入）
    def safe_float(x):
        try:
            if x == "" or x is None:
                return 0.0
            return float(x)
        except Exception:
            return 0.0

    v4 = safe_float(row[3]) if len(row) > 3 else 0.0
    v28 = safe_float(row[27]) if len(row) > 27 else 0.0
    v29 = safe_float(row[28]) if len(row) > 28 else 0.0
    v33 = safe_float(row[32]) if len(row) > 32 else 0.0
    v34 = safe_float(row[33]) if len(row) > 33 else 0.0

    row.append(round(v4 * v29, 1))  # 4番目×29番目
    row.append(round(v4 * v28, 1))  # 4番目×28番目
    row.append(round(v4 * v33, 1))  # 4番目×33番目
    row.append(round(v4 * v34, 1))  # 4番目×34番目

    write_to_excel([row], file_name)



# 共有バッファとロック
shared_data = {"logger1": None, "logger2": None}
shared_lock = threading.Lock()
write_event = threading.Event()

def process_and_save_data_sync(logger_id, address_list, value_list, file_name):
    with shared_lock:
        now = datetime.now()
        timestamp = now.strftime("%Y-%m-%d %H:%M:%S")
        shared_data[logger_id] = (timestamp, address_list, value_list, file_name)
        # Loggerが2つある場合のみ両方揃ったら同時書き込み
        if (ser_logger1 and ser_logger2) and shared_data["logger1"] and shared_data["logger2"]:
            t, addr1, val1, file1 = shared_data["logger1"]
            _, addr2, val2, file2 = shared_data["logger2"]
            process_and_save_data(timestamp, addr1, val1, file1)
            process_and_save_data(timestamp, addr2, val2, file2)
            shared_data["logger1"] = None
            shared_data["logger2"] = None
        # Loggerが1つだけの場合は即書き込み
        elif (ser_logger1 and not ser_logger2 and logger_id == "logger1") or (ser_logger2 and not ser_logger1 and logger_id == "logger2"):
            t, addr, val, file = shared_data[logger_id]
            process_and_save_data(timestamp, addr, val, file)
            shared_data[logger_id] = None

# --- handle_logger1_inverter1 の該当部分を修正 ---
def handle_logger1_inverter1():
    """ Logger1とInverter1間のデータ送受信を高速に処理するスレッド """
    try:
        data_buffer = {}
        inverter1_frame_buffer = b''  # 受信フレーム結合用バッファ
        while True:
            did_process = False
            if ser_logger1 and ser_inverter1:
                if ser_logger1.in_waiting:
                    logger1_data = ser_logger1.read(ser_logger1.in_waiting)
                    #print(f"[Logger1→Inverter1] 受信: {logger1_data.hex()}")
                    reg_address = extract_modbus_register_address(logger1_data)
                    ser_inverter1.write(logger1_data)
                    #print(f"[Logger1→Inverter1] 送信: {logger1_data.hex()}")
                    did_process = True

                if ser_inverter1.in_waiting:
                    inverter1_data = ser_inverter1.read(ser_inverter1.in_waiting)
                    #print(f"[Inverter1→Logger1] 受信: {inverter1_data.hex()}")
                    # 受信したら即Loggerへ送信
                    ser_logger1.write(inverter1_data)
                    #print(f"[Inverter1→Logger1] 送信: {inverter1_data.hex()}")
                    # フレームをバッファに追加
                    inverter1_frame_buffer += inverter1_data
                    # フレーム終端判定（例：CRC2バイトを含む長さで判定、またはプロトコル仕様に応じて調整）
                    # ここでは「連続で2つ以上のフレームが来たら結合して処理」とする
                    # 例: 1フレーム目受信後、2フレーム目受信時に結合して処理
                    # CRC2バイトを含むフレーム長を判定する場合は下記のように調整
                    # ここでは単純に2フレーム分たまったら処理
                    if len(inverter1_frame_buffer) > 0:
                        # 例: 2フレーム分たまったら処理（必要に応じて判定方法を調整）
                        # ここでは100ms待って追加受信がなければ処理する簡易方式
                        time.sleep(0.1)
                        if not ser_inverter1.in_waiting:
                            # ファイル書き込み処理は結合したフレームで
                            reg_value = extract_modbus_register_value(inverter1_frame_buffer)
                            reg_address = extract_modbus_register_address(logger1_data)
                            if reg_value and reg_address:
                                addr_ints = [int(a, 16) for a in reg_address]
                                for addr, val in zip(addr_ints, reg_value):
                                    data_buffer[addr] = val
                                if 0xf040 in addr_ints:
                                    write_addrs = [addr for addr in modbus_addr_sorted if 0x000b <= addr <= 0xf04b]
                                    row_values = [data_buffer.get(addr, "") for addr in write_addrs]
                                    file_name = f"test1-{datetime.now().strftime('%Y-%m-%d')}.xlsx"
                                    process_and_save_data_sync("logger1", write_addrs, row_values, file_name)
                                    data_buffer.clear()
                            inverter1_frame_buffer = b''  # バッファクリア
                    did_process = True

            if not did_process:
                time.sleep(0.001)
    except Exception as e:
        print(f"Logger1-Inverter1スレッドでエラーが発生しました: {e}")

def handle_logger2_inverter2():
    """ Logger2とInverter2間のデータ送受信を高速に処理するスレッド """
    try:
        data_buffer = {}
        inverter2_frame_buffer = b''
        while True:
            did_process = False
            if ser_logger2 and ser_inverter2:
                if ser_logger2.in_waiting:
                    logger2_data = ser_logger2.read(ser_logger2.in_waiting)
                    print(f"[Logger2→Inverter2] 受信: {logger2_data.hex()}")
                    reg_address = extract_modbus_register_address(logger2_data)
                    ser_inverter2.write(logger2_data)
                    print(f"[Logger2→Inverter2] 送信: {logger2_data.hex()}")
                    did_process = True

                if ser_inverter2.in_waiting:
                    inverter2_data = ser_inverter2.read(ser_inverter2.in_waiting)
                    print(f"[Inverter2→Logger2] 受信: {inverter2_data.hex()}")
                    ser_logger2.write(inverter2_data)
                    print(f"[Inverter2→Logger2] 送信: {inverter2_data.hex()}")
                    inverter2_frame_buffer += inverter2_data
                    time.sleep(0.1)
                    if not ser_inverter2.in_waiting:
                        reg_value = extract_modbus_register_value(inverter2_frame_buffer)
                        reg_address = extract_modbus_register_address(logger2_data)
                        if reg_value and reg_address:
                            addr_ints = [int(a, 16) for a in reg_address]
                            for addr, val in zip(addr_ints, reg_value):
                                data_buffer[addr] = val
                            if 0xf040 in addr_ints:
                                write_addrs = [addr for addr in modbus_addr_sorted if 0x000b <= addr <= 0xf04b]
                                row_values = [data_buffer.get(addr, "") for addr in write_addrs]
                                file_name = f"test2-{datetime.now().strftime('%Y-%m-%d')}.xlsx"
                                process_and_save_data_sync("logger2", write_addrs, row_values, file_name)
                                data_buffer.clear()
                        inverter2_frame_buffer = b''
                    did_process = True

            if not did_process:
                time.sleep(0.001)
    except Exception as e:
        print(f"Logger2-Inverter2スレッドでエラーが発生しました: {e}")

# スレッドの作成
threads = []

if ser_logger1 and ser_inverter1:
    thread1 = threading.Thread(target=handle_logger1_inverter1, daemon=True)
    threads.append(thread1)

if ser_logger2 and ser_inverter2:
    thread2 = threading.Thread(target=handle_logger2_inverter2, daemon=True)
    threads.append(thread2)

# スレッドの開始
for thread in threads:
    thread.start()

# メインスレッドでスレッドの終了を待機
try:
    for thread in threads:
        thread.join()
except KeyboardInterrupt:
    print("データ交換を終了します。")
finally:
    if ser_logger1 and ser_logger1.is_open:
        ser_logger1.close()
    if ser_inverter1 and ser_inverter1.is_open:
        ser_inverter1.close()
    if ser_logger2 and ser_logger2.is_open:
        ser_logger2.close()
    if ser_inverter2 and ser_inverter2.is_open:
        ser_inverter2.close()

