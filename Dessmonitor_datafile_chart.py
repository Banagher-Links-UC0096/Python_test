import csv
import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook
import matplotlib.pyplot as plt
import matplotlib.font_manager as fm
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import seaborn as sns
import numpy as np
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler
import os


# 日本語フォント設定
font_path = 'c:/windows/Fonts/meiryo.ttc'
jp_font = fm.FontProperties(fname=font_path)

# ファイル選択関数
def select_file(title):
    root = tk.Tk()
    root.withdraw()
    return filedialog.askopenfilename(title=title, filetypes=[('CSV or Excel files', '*.csv;*.xlsx')])

# CSVデータ取得関数
def extract_data_csv(filepath, col, is_string=False):
    data = []
    with open(filepath, mode='r', encoding='shift_jis') as file:
        reader = csv.reader(file)
        for row in reader:
            try:
                value = row[col - 1]  # 0-indexed
                if is_string:
                    data.append(value)
                else:
                    # 余分な文字を取り除く処理
                    value = value.strip("[]")  # "["や"]"を削除
                    data.append(int(float(value)))  # 数値に変換
            except ValueError:
                print(f"数値に変換できないデータをスキップしました: {value}")
            except IndexError:
                print("指定された列番号が範囲外です")
    return data

# データ取得関数
def extract_data(sheet, col, is_string=False):
    data = []
    for row in sheet.iter_rows(min_row=1, min_col=col, max_col=col):
        for cell in row:
            if cell.value is not None:  # 空セルの処理を防ぐ
                if is_string:  # 文字列として処理
                    data.append(str(cell.value))
                else:  # 数値として処理
                    try:
                        data.append(int(float(cell.value)))  # 数値に変換
                    except ValueError:
                        print(f"数値に変換できないデータをスキップしました: {cell.value}")
    return data

# ファイル検索関数（後半10文字一致かつ異なるファイルを探す）
def find_matching_file(directory, base_filename_tail, file1_path):
    for file in os.listdir(directory):
        file_path = os.path.join(directory, file)  # フルパスを作成
        # ファイル名の後半10文字が一致し、かつ1つ目のファイルと異なるファイルを選択
        if file.endswith(base_filename_tail) and os.path.abspath(file_path) != os.path.abspath(file1_path):
            return file_path
    return None  # 一致するファイルが見つからない場合

# グラフ描画関数
def plot_graph(x, y_data, labels, colors):
    for y, label, color in zip(y_data, labels, colors):
        plt.plot(x, y, label=label, linestyle='-', color=color, linewidth=0.2 if 'Total' not in label else 1)

#def toggle_line_visibility(line, button):
#    line.set_visible(not line.get_visible())  # ラインの表示・非表示を切り替え
#    button.config(text=f"{line.get_label()} {'表示' if line.get_visible() else '非表示'}")
#    canvas.draw()

# ファイル読み込み
file1_path = select_file("Dessmonitor Detail report Download File1 or today_logfile.csv")
if file1_path.endswith('.csv'):
    # CSV処理
    # 抽出対象の列番号とオプション (is_string=True/False)
    columns = [
        (1, True),  # data0: 1列目 (文字列として処理)
        (19, False), (62, False), # data1 : 18列目 data2 : 44+18列目
        (9 , False), (52, False), # data4 :  8列目 data5 : 44+ 8列目
        (15, False), (47, False), # data7 :  3列目 data8 : 44+ 3列目
        (16, False), (48, False), # data10:  4列目 data11: 44+ 4列目
        (2 , False), (52, False), # data13:  8列目 data14: 44+ 8列目
        (3 , False), (53, False)] # data16:  9列目 data17: 44+ 9列目
    # ループで一括処理
    extracted_data = [extract_data_csv(file1_path, col, is_string=is_str) for col, is_str in columns]
    # 結果を個別変数に戻す
    data0, data1, data2, data4, data5, data7, data8, data10, data11, data13, data14, data16, data17 = extracted_data
else:
    if file1_path:  # 2つ目のファイルを自動検索
        directory = os.path.dirname(file1_path)  # 1つ目のファイルのディレクトリを取得
        file1_name_tail = os.path.basename(file1_path)[-10:]  # ファイル名の後半10文字を取得
        file2_path = find_matching_file(directory, file1_name_tail, file1_path)  # 修正済みの関数を使用
    if file2_path:
        print(f"2つ目のファイルが見つかりました: {file2_path}")
    else:
        print("2つ目のファイルが見つかりませんでした")   
    # Excel処理
    sheet1 = load_workbook(file1_path).active
    sheet2 = load_workbook(file2_path).active

    # 抽出対象の列番号と対応するシート
    columns = [
        (sheet1, 1 , True),  # data0: 1列目 (文字列として処理)
        (sheet1, 15, False), (sheet2, 15, False), # data1 :PV入力 14列目 data2 : 14列目
        (sheet1, 22, False), (sheet2, 22, False), # data4 :AC出力 21列目 data5 : 21列目
        (sheet1, 4 , False), (sheet2, 4 , False), # data7 :BT電圧  3列目 data8 :  3列目
        (sheet1, 5 , False), (sheet2, 5 , False), # data10:BT電流  4列目 data11:  4列目
        (sheet1, 9 , False), (sheet2, 9 , False), # data13:AC電圧  8列目 data14:  8列目
        (sheet1, 10, False), (sheet2, 10, False)] # data16:AC電流  9列目 data17:  9列目
    # ループでデータ抽出
    extracted_data = [extract_data(sheet, col, is_string=is_str) for sheet, col, is_str in columns]
    # 結果を個別変数に戻す
    data0, data1, data2, data4, data5, data7, data8, data10, data11, data13, data14, data16, data17 = extracted_data
# データ調整
min_length = min(map(len, [data0, data1, data2, data4, data5,data7,data8,data10,data11,data13,data14,data16,data17]))                    
data0, data1, data2, data4, data5,data7,data8,data10,data11,data13,data14,data16,data17 = [d[:min_length] for d in [
    data0, data1, data2, data4, data5,data7,data8,data10,data11,data13,data14,data16,data17]]
data3 , data6 = [a + b for a, b in zip(data1 , data2 )],[a + b for a, b in zip(data4 , data5 )]
data9 , data12= [a * b for a, b in zip(data7 , data10)],[a * b for a, b in zip(data8 , data11)]
data15, data18= [a * b for a, b in zip(data13, data16)],[a * b for a, b in zip(data14, data17)]
data19, data20= [a + b for a, b in zip(data9 , data12)],[a + b for a, b in zip(data15, data18)]
#for a in range(len(data1)):data1[a], data2[a], data3[a]= -data1[a], -data2[a], -data3[a] # PV_power反転
# 対象リストを一括で反転
lists = [data0, data1, data2, data3, data4, data5, data6, data9, data12, data19, data15, data18, data20]
lists = [lst[::-1] for lst in lists]
# 反転したリストを個別に戻したい場合
data0, data1, data2, data3, data4, data5, data6, data9, data12, data19, data15, data18, data20 = lists

# グラフ描画設定
sns.set(style='darkgrid', palette='winter_r')
plt.figure(figsize=(15, 8))
plt.xticks(rotation=80)
plt.yticks(rotation=-10)

# プロット
plot_graph(data0, [data1, data2, data3, data4, data5, data6,data9,data12,data19,data15,data18,data20],
           ['PV1_power', 'PV2_power', 'PV_Total_power', 'Inv1_power', 'Inv2_power', 'INV_Total_power','Batt_power1','Batt_power2','Batt_Total_power','Grid_power1','Grid_power2','Grid_Total_power'],
           ['red', 'red', 'red', 'blue', 'blue', 'blue','green','green','green','purple','purple','purple'])

# 最大値マーク
for data, label, color,ha in [(data3, "Max_PV_power", "red","right"), (data6, "Max_INV_power", "blue","left"),(data19, "Max_Batt_power", "green","right"),(data20, "Max_Grid_power", "purple","right")]:
    max_idx = np.argmax(data)
    plt.scatter(data0[max_idx], data[max_idx], color=color, label=f"{label}")
    plt.text(data0[max_idx], data[max_idx], f"({data0[max_idx]}, {data[max_idx]}W)", fontsize=10, color=color, ha=ha)

# Batt_powerデータの最低値マーク
for data, label, color in [(data19, "Min_Batt_Total_power", "green")]:
    min_idx = np.argmin(data)  # 最低値のインデックスを取得
    plt.scatter(data0[min_idx], data[min_idx], color=color, marker='x', label=f"{label}")  # 最低値マーク
    plt.text(data0[min_idx], data[min_idx], f"({data0[min_idx]}, {data[min_idx]}W)", fontsize=10, color=color, ha="left")  # 最低値にテキスト表示

# グラフ設定
plt.title('電力量', fontsize=40, fontproperties=jp_font)
plt.xlabel('日時', fontsize=12, fontproperties=jp_font)
plt.ylabel('電力', fontsize=20, fontproperties=jp_font)
plt.legend(bbox_to_anchor=(1, 1), loc='upper left')
plt.grid(True)
plt.gca().xaxis.set_major_locator(plt.MultipleLocator(10))
plt.gca().yaxis.set_major_locator(plt.MultipleLocator(500))

# Maximize and set window title
manager = plt.get_current_fig_manager()
manager.window.state('zoomed')  # For maximizing (Windows)
manager.set_window_title("PV & Inverter Power Chart")  # Set window title
# Show the plot

plt.show()

