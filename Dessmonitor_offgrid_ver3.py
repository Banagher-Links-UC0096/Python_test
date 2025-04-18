import openpyxl
from datetime import datetime, date
from tkinter import Tk, filedialog
from collections import defaultdict
import os
import jpholiday  # 日本の祝日判定ライブラリ

# フォルダー選択ダイアログを表示
def select_folder():
    root = Tk()
    root.withdraw()  # Tkinterウィンドウを非表示にする
    folder_path = filedialog.askdirectory(
        title="フォルダーを選択してください"
    )
    if not folder_path:
        print("フォルダーが選択されませんでした。")
        exit()
    print(f"選択されたフォルダー: {folder_path}")
    return folder_path

# 時間帯ごとのデータを計算する関数
def calculate_time_data(file_path):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    # 時間帯ごとの最終データを格納する辞書
    time_ranges = {
        "0時～7時": (0, 7),
        "7時～9時": (7, 9),
        "9時～17時": (9, 17),
        "17時～23時": (17, 23),
        "23時～24時": (23, 24)
    }
    time_last_values = defaultdict(float)
    loss_last_values = defaultdict(float)  # 損失電力の最終データ

    # データを逆順で処理（行の終わりが先頭データ）
    rows = list(sheet.iter_rows(min_row=2))  # ヘッダーをスキップ
    for row in reversed(rows):  # 行を逆順に処理
        date_cell = row[0].value  # 1列目の日付データ

        # 列数チェック（31列目と32列目が存在するか確認）
        if len(row) <= 31:
            continue

        value_cell = row[30].value  # 31列目のデータ
        loss_cell = row[31].value  # 32列目のデータ

        # date_cell の確認と変換
        if isinstance(date_cell, str):
            try:
                date_cell = datetime.strptime(date_cell, "%Y-%m-%d %H:%M:%S")
            except ValueError:
                continue
        elif not isinstance(date_cell, datetime):
            continue

        # value_cell の確認と変換
        try:
            value_cell = float(value_cell)
        except (ValueError, TypeError):
            continue

        # loss_cell の確認と変換
        try:
            if isinstance(loss_cell, str):
                loss_cell = loss_cell.replace(",", "")  # カンマを削除
            loss_cell = float(loss_cell)  # 数値に変換
        except (ValueError, TypeError):
            continue

        # 時間部分を抽出
        hour = date_cell.hour

        # 土日祝の判定
        is_weekend_or_holiday = date_cell.weekday() >= 5 or jpholiday.is_holiday(date_cell.date())

        # 時間帯ごとの最終データを更新
        for label, (start, end) in time_ranges.items():
            if start <= hour < end:
                if label == "9時～17時" and is_weekend_or_holiday:
                    # 土日祝の9時～17時は「ホームタイム」に分類
                    time_last_values["ホームタイム"] = value_cell
                    loss_last_values["ホームタイム"] = loss_cell
                else:
                    time_last_values[label] = value_cell
                    loss_last_values[label] = loss_cell

    # 時間帯ごとの増加量を計算
    time_increments = defaultdict(float)
    loss_increments = defaultdict(float)
    previous_value = None
    previous_loss = None
    for label in time_ranges.keys():
        if previous_value is not None:
            time_increments[label] = time_last_values[label] - previous_value
        else:
            time_increments[label] = time_last_values[label]  # 初回はそのまま
        previous_value = time_last_values[label]

        if previous_loss is not None:
            loss_increments[label] = loss_last_values[label] - previous_loss
        else:
            loss_increments[label] = loss_last_values[label]  # 初回はそのまま
        previous_loss = loss_last_values[label]

    # タイムゾーンごとの集計
    time_zone_totals = {
        "ナイトタイム": time_increments["0時～7時"] + time_increments["23時～24時"],
        "ホームタイム": time_increments["7時～9時"] + time_increments["17時～23時"] + time_increments.get("ホームタイム", 0),
        "デイタイム": time_increments["9時～17時"] if not is_weekend_or_holiday else 0
    }
    loss_zone_totals = {
        "ナイトタイム": loss_increments["0時～7時"] + loss_increments["23時～24時"],
        "ホームタイム": loss_increments["7時～9時"] + loss_increments["17時～23時"] + loss_increments.get("ホームタイム", 0),
        "デイタイム": loss_increments["9時～17時"] if not is_weekend_or_holiday else 0
    }

    return time_zone_totals, loss_zone_totals

# メイン処理
folder_path = select_folder()

# 選択したフォルダー内のすべてのExcelファイルを検索
all_files = []
for root, _, files in os.walk(folder_path):  # フォルダー内を再帰的に探索
    for file in files:
        if file.endswith(".xlsx") or file.endswith(".xlsm"):
            all_files.append(os.path.join(root, file))

# 処理対象のファイルを表示
print("処理対象のファイル:")
for f in all_files:
    print(f"  - {f}")

# 燃料費調整単価と再エネ賦課金単価を設定
fuel_adjustment_rates = {
    "2023-04": 2.93, 
    "2023-05": 1.95,
    "2023-06": 0.60,
    "2023-07":-0.94,
    "2023-08":-2.57,
    "2023-09":-3.74,
    "2023-10":-0.73,
    "2023-11":-0.96,
    "2023-12":-1.10,
    "2024-01":-1.01,
    "2024-02":-0.82,
    "2024-03":-0.31,
    "2024-04":-0.10,
    "2024-05": 0.04,
    "2024-06": 1.51,
    "2024-07": 2.84,
    "2024-08": 2.54,
    "2024-09":-1.55,
    "2024-10":-1.25,
    "2024-11": 0.30,
    "2024-12": 2.59,
    "2025-01": 2.33,
    "2025-02":-0.15,
    "2025-03": 0.06,
    "2025-04": 1.64,
    "2025-05": 2.84,  # 例: 5月以降の単価
    # 必要に応じて追加
}

renewable_energy_rates = {
    "2022": 3.45,  # 例: 2022年度の単価
    "2023": 1.40,  # 例: 2023年度の単価
    "2024": 3.49,  # 例: 2024年度の単価
    "2025": 3.98,  # 例: 2025年度の単価
    # 必要に応じて追加
}

# 単価を取得する関数
def get_fuel_adjustment_rate(year_month):
    return fuel_adjustment_rates.get(year_month, 0.0)  # デフォルト値は0.0

# 再エネ賦課金単価を取得する関数
def get_renewable_energy_rate(year, month):
    # 1～4月は前年の単価を適用
    if month in [1, 2, 3, 4]:
        previous_year = str(int(year) - 1)
        return renewable_energy_rates.get(previous_year, 0.0)
    # 5月以降は当年の単価を適用
    return renewable_energy_rates.get(year, 0.0)

# 全ファイルのデータを計算
total_time_zone_totals = defaultdict(float)  # タイムゾーンごとの合計を保持
total_loss_zone_totals = defaultdict(float)  # 損失電力の合計を保持

# データ期間を記録するための変数
all_dates = []

for file in all_files:
    # 各ファイルのデータを計算
    time_zone_totals, loss_zone_totals = calculate_time_data(file)

    # データ期間を取得（1列目の日付データから算出）
    wb = openpyxl.load_workbook(file)
    sheet = wb.active
    file_dates = []  # ファイルごとの日付を記録
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):  # ヘッダーをスキップ
        date_cell = row[0].value

        # 日付データが文字列の場合を考慮して変換
        if isinstance(date_cell, str):
            try:
                date_cell = datetime.strptime(date_cell, "%Y-%m-%d %H:%M:%S")
            except ValueError:
                try:
                    date_cell = datetime.strptime(date_cell, "%Y-%m-%d")
                except ValueError:
                    print(f"無効な日付データ: {date_cell}")
                    continue

        # 日付データが datetime 型の場合のみ処理
        if isinstance(date_cell, datetime):
            file_dates.append(date_cell.date())
            all_dates.append(date_cell.date())

    # ファイルごとのデータ期間を取得
    if file_dates:
        file_start_date = min(file_dates)
        file_end_date = max(file_dates)
        #print(f"\nファイル: {file}")
        #print(f"データ期間: {file_start_date} ～ {file_end_date}")

        # 適用単価を取得
        year_month = file_start_date.strftime("%Y-%m")  # ファイルの開始年月
        year = file_start_date.strftime("%Y")  # ファイルの開始年
        month = file_start_date.month  # ファイルの開始月
        fuel_adjustment_cost_per_kwh = get_fuel_adjustment_rate(year_month)
        renewable_energy_cost_per_kwh = get_renewable_energy_rate(year, month)

        #print("\n適用単価:")
        #print(f"燃料費調整単価 ({year_month}): {fuel_adjustment_cost_per_kwh:.2f} 円/kWh")
        #print(f"再エネ賦課金単価 ({year}): {renewable_energy_cost_per_kwh:.2f} 円/kWh")

    # 各タイムゾーンのデータを集計
    for zone, total in time_zone_totals.items():
        total_time_zone_totals[zone] += total
    for zone, loss in loss_zone_totals.items():
        total_loss_zone_totals[zone] += loss

# 全体のデータ期間を表示
if all_dates:
    start_date = min(all_dates)  # 最小の日付
    end_date = max(all_dates)    # 最大の日付
    print(f"\n全体のデータ期間: {start_date} ～ {end_date}")

    # データ期間に基づいて適用した単価を表示
    start_year_month = start_date.strftime("%Y-%m")
    end_year_month = end_date.strftime("%Y-%m")
    start_year = start_date.year
    end_year = end_date.year

    print("\n適用した燃料費調整単価:")
    for year_month, rate in fuel_adjustment_rates.items():
        if start_year_month <= year_month <= end_year_month:
            print(f"{year_month}: {rate:.2f} 円/kWh")

    print("\n適用した再エネ賦課金単価:")
    for year, rate in renewable_energy_rates.items():
        if start_year <= int(year) <= end_year:
            print(f"{year}: {rate:.2f} 円/kWh")
else:
    print("\nデータ期間: データがありません")

# 適用した燃料費調整単価と再エネ賦課金単価を表示
#print("\n適用した単価:")
#print(f"燃料費調整単価: {fuel_adjustment_cost_per_kwh:.2f} 円/kWh")
#print(f"再エネ賦課金単価: {renewable_energy_cost_per_kwh:.2f} 円/kWh")

# 各タイムゾーンの集計を表示
print("\nタイムゾーンごとの集計:")
total_cost = 0
rates = {"デイタイム": 34.06, "ホームタイム": 26.00, "ナイトタイム": 16.11}
total_renewable_energy_cost = 0
total_fuel_adjustment_cost = 0

for zone, total in total_time_zone_totals.items():
    # 各単価を適用して計算
    cost = total * rates[zone]
    renewable_energy_cost = total * renewable_energy_cost_per_kwh
    fuel_adjustment_cost = total * fuel_adjustment_cost_per_kwh

    total_cost += cost
    total_renewable_energy_cost += renewable_energy_cost
    total_fuel_adjustment_cost += fuel_adjustment_cost

    # 計算結果のみを表示
    print(f"{zone}: {total:.2f} kWh, 金額: {cost:.2f} 円, "
          f"再エネ賦課金: {renewable_energy_cost:.2f} 円, 燃料費調整額: {fuel_adjustment_cost:.2f} 円")

# 損失電力の集計を表示
print("\n損失電力の集計:")
total_loss_cost = 0
total_loss_renewable_cost = 0
total_loss_fuel_adjustment_cost = 0

for zone, loss in total_loss_zone_totals.items():
    # 各単価を適用して計算
    loss_cost = loss * rates[zone]
    loss_renewable_cost = loss * renewable_energy_cost_per_kwh
    loss_fuel_adjustment_cost = loss * fuel_adjustment_cost_per_kwh

    total_loss_cost += loss_cost
    total_loss_renewable_cost += loss_renewable_cost
    total_loss_fuel_adjustment_cost += loss_fuel_adjustment_cost

    # 計算結果のみを表示
    print(f"{zone}: {loss:.2f} kWh, 損失金額: {loss_cost:.2f} 円, "
          f"再エネ賦課金: {loss_renewable_cost:.2f} 円, 燃料費調整額: {loss_fuel_adjustment_cost:.2f} 円")

# 再エネ賦課金単価と燃料費調整単価を計算
total_energy = sum(total_time_zone_totals.values())

print(f"\n再エネ賦課金: {total_renewable_energy_cost:.2f} 円")
print(f"燃料費調整額: {total_fuel_adjustment_cost:.2f} 円")

# 合計金額を表示
total_cost += total_renewable_energy_cost + total_fuel_adjustment_cost
print(f"\n太陽光発電による節約金額: {total_cost:.2f} 円")

# 損失を含む合計金額を計算
total_loss_adjustment = total_loss_cost + total_loss_renewable_cost + total_loss_fuel_adjustment_cost
adjusted_total_cost = total_cost - total_loss_adjustment
print(f"損失電力を含む経済効果: {adjusted_total_cost:.2f} 円")
